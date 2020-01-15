from pyrfc import Connection, ABAPApplicationError, ABAPRuntimeError, LogonError, CommunicationError
import pandas as pd
from pprint import PrettyPrinter
import re
import os
from datetime import date
from sharepoint_get_file import download_file, upload_file
import openpyxl



class RfcConnection:
    def __init__(self):
        self.conn = None
        while self.conn is None:
            try:
                self.conn = Connection(user=df_config.iloc[6,1], passwd=df_config.iloc[7,1],
                    mshost=df_config.iloc[0,1],
                    mserv=df_config.iloc[1,1],
                    sysid=df_config.iloc[5,1],
                    sysnr=df_config.iloc[2,1],
                    group=df_config.iloc[4,1],
                    client=df_config.iloc[3,1],
                    lang=df_config.iloc[8,1])

            except CommunicationError:
                print ("Could not connect to server.")
                raise
            except LogonError:
                print ("Could not log in. Wrong credentials?")
                raise
            except (ABAPApplicationError, ABAPRuntimeError):
                print ("An error occurred.")
                raise


    def rfc_get_system_info(self):
        return self.conn.call("RFC_GET_SYSTEM_INFO")


    def get_report(self, group, report, variant, file_name):
        print("gathering data from "+report+" ...")
        data = self.conn.call("RSAQ_REMOTE_QUERY_CALL", USERGROUP=group, QUERY=report, VARIANT=variant,
                                DATA_TO_MEMORY="X", EXTERNAL_PRESENTATION="Z")
        columns = []

        for i in data['LISTDESC']:
            if i['LID'] == 'G00':
                columns.append(i['FCOL'])

        datastring = ""
        datalist = [list(x.values()) for x in data['LDATA']]
        datastring = datastring.join(str(r) for v in datalist for r in v)
        datastring = datastring.split(';/')[0]
        datastring = datastring.split(';')
        datalist = []

        for i in datastring:
            datalist.append(re.split(',\d\d\d:', i[4:]))

        results = pd.DataFrame(data=datalist, columns=columns)

        # Forcing PATH for to_excel function
        dir_path = os.path.dirname(os.path.realpath(__file__))
        file_name = file_name+'.xlsx'

        results.to_excel(os.path.join(dir_path, file_name), index=False)

        # THEN SEND THE FILE TO SHAREPOINT
        upload_file(file_name)
        
        print("ETL ok")

        return results

if __name__ == "__main__":
    # TEST EXCEL FILE
    df_config = download_file()

    tmp_conn = RfcConnection()



    excel_file = openpyxl.load_workbook('Reporting_ETL.xlsx')
    ws = excel_file['Report']

    for i in range(5,11):
        Action = ws.cell(row = i, column = 1).value
        Type = ws.cell(row = i, column = 3).value
        Schedule = ws.cell(row = i, column = 9).value
        GROUP = ws.cell(row = i, column = 4).value
        REPORT = ws.cell(row = i, column = 5).value
        VARIANT = ws.cell(row = i, column = 6).value
        Filename = ws.cell(row = i, column = 7).value


        if Action == "YES" and Type == "REPORT":
            print(VARIANT)
            tmp_conn.get_report(GROUP, REPORT, VARIANT, Filename)
            ws.cell(row = i, column = 10).value = date.today()
        else:
            pass
    
    excel_file.save(filename="Reporting_ETL.xlsx")
    filename = "Reporting_ETL.xlsx"
    upload_file(filename)






